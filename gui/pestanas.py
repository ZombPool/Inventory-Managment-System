import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
import logging
from typing import Dict, Optional, Tuple, List
import os

# --- Basic Logging Configuration ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='almacen.log',
    filemode='a' # Append to log file
)
logger = logging.getLogger(__name__)

# --- ExcelManager Class ---
class ExcelManager:
    """Class to handle all optimized Excel operations."""

    def __init__(self, archivo_excel: str):
        self.archivo_excel = archivo_excel
        self._wb = None
        self._cache = {}
        self.column_mapping = {
            'Ingresos de almacén': {
                'Fecha': 'A', 'N° de parte': 'B', 'Nombre': 'C',
                'Descripción': 'D', 'Unidad': 'F', 'Cantidad': 'G',
                'Almacén': 'H', 'Ubicación': 'I', 'Encargado': 'J',
                'Comentarios': 'K'
            },
            'Salidas de almacén': {
                'Fecha': 'A', 'N° de parte': 'B', 'Nombre': 'C',
                'Descripción': 'D', 'Unidad': 'F', 'Cantidad': 'G',
                'Almacén': 'H', 'Ubicación': 'I', 'Encargado': 'J',
                'Comentarios': 'K'
            },
            'Control de inventarios': {
                'N° de parte': 'A', 'Nombre': 'B', 'Stock actual': 'C',
                'Stock mínimo': 'D', 'Stock máximo': 'E', 'Estado': 'F'
            }
        }
        self._ensure_sheets_exist() # Ensure sheets are present on initialization

    @property
    def workbook(self):
        """Property that handles lazy loading of the workbook."""
        if self._wb is None:
            try:
                self._wb = load_workbook(self.archivo_excel)
                logger.info("Workbook loaded successfully.")
            except FileNotFoundError:
                logger.error(f"Error: Excel file not found at {self.archivo_excel}")
                messagebox.showerror("Error de Archivo", f"El archivo Excel no se encontró en:\n{self.archivo_excel}\nPor favor, verifique la ruta.")
                raise
            except Exception as e:
                logger.error(f"Error loading workbook: {str(e)}")
                messagebox.showerror("Error al Cargar", f"Error al cargar el archivo Excel: {e}")
                raise
        return self._wb

    def _ensure_sheets_exist(self):
        """Ensures all required sheets exist in the workbook."""
        try:
            wb = self.workbook # This will load the workbook
            for sheet_name in self.column_mapping.keys():
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                    # Add headers if sheet is new (adjust based on your actual headers)
                    if sheet_name == 'Ingresos de almacén':
                        ws = wb[sheet_name]
                        ws['A1'] = "REGISTRO DE INGRESOS EN ALMACÉN"
                        ws['A2'] = "Fecha"
                        ws['B2'] = "N° de parte"
                        ws['C2'] = "Nombre"
                        ws['D2'] = "Descripción"
                        ws['E2'] = " " # Added a placeholder for E2 if F is used for Unit
                        ws['F2'] = "Unidad"
                        ws['G2'] = "Cantidad"
                        ws['H2'] = "Almacén"
                        ws['I2'] = "Ubicación"
                        ws['J2'] = "Encargado"
                        ws['K2'] = "Comentarios"
                    elif sheet_name == 'Salidas de almacén':
                        ws = wb[sheet_name]
                        ws['A1'] = "REGISTRO DE SALIDAS EN ALMACÉN"
                        ws['A2'] = "Fecha"
                        ws['B2'] = "N° de parte"
                        ws['C2'] = "Nombre"
                        ws['D2'] = "Descripción"
                        ws['E2'] = " " # Added a placeholder for E2 if F is used for Unit
                        ws['F2'] = "Unidad"
                        ws['G2'] = "Cantidad"
                        ws['H2'] = "Almacén"
                        ws['I2'] = "Ubicación"
                        ws['J2'] = "Encargado"
                        ws['K2'] = "Comentarios"
                    elif sheet_name == 'Control de inventarios':
                        ws = wb[sheet_name]
                        ws['A1'] = "CONTROL DE INVENTARIOS"
                        ws['A2'] = "N° de parte"
                        ws['B2'] = "Nombre"
                        ws['C2'] = "Stock actual"
                        ws['D2'] = "Stock mínimo"
                        ws['E2'] = "Stock máximo"
                        ws['F2'] = "Estado"
                    logger.warning(f"Sheet '{sheet_name}' was missing and has been created.")
            self.save() # Save after creating missing sheets
        except Exception as e:
            logger.error(f"Error ensuring sheets exist: {str(e)}")
            messagebox.showerror("Error de Hojas", f"No se pudieron asegurar las hojas del Excel: {e}")

    def get_sheet(self, sheet_name: str):
        """Obtains a specific sheet with caching."""
        if sheet_name not in self._cache:
            try:
                self._cache[sheet_name] = self.workbook[sheet_name]
                logger.debug(f"Sheet '{sheet_name}' loaded into cache.")
            except KeyError:
                logger.error(f"Sheet '{sheet_name}' not found.")
                messagebox.showerror("Error de Hoja", f"La hoja '{sheet_name}' no se encontró en el archivo Excel.")
                raise
        return self._cache[sheet_name]

    def save(self):
        """Saves changes to the Excel file."""
        try:
            if self._wb: # Ensure workbook is loaded before saving
                self._wb.save(self.archivo_excel)
                logger.info("Changes saved successfully.")
            else:
                logger.warning("Attempted to save but workbook was not loaded.")
        except Exception as e:
            logger.error(f"Error saving: {str(e)}")
            messagebox.showerror("Error al Guardar", f"Error al guardar los cambios en Excel: {e}")
            raise

    def find_part(self, sheet_name: str, part_number: str) -> Optional[int]:
        """Searches for a part number and returns the row if it exists.
        Starts search from row 3 (after headers)."""
        ws = self.get_sheet(sheet_name)
        # Assuming part numbers are in column B for Ingresos/Salidas, and A for Control
        col_letter = self.column_mapping[sheet_name]['N° de parte']
        col_idx = ord(col_letter) - ord('A') + 1

        # Iterate from the first data row (row 3)
        for row_idx in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip() == str(part_number).strip():
                logger.debug(f"Part '{part_number}' found in row {row_idx} of '{sheet_name}'.")
                return row_idx
        logger.debug(f"Part '{part_number}' not found in '{sheet_name}'.")
        return None

    def get_cell_value(self, sheet_name: str, row: int, column_letter: str):
        """Gets the value of a specific cell."""
        ws = self.get_sheet(sheet_name)
        return ws[f'{column_letter}{row}'].value

    def update_cell(self, sheet_name: str, row: int, column_letter: str, value):
        """Updates the value of a specific cell."""
        ws = self.get_sheet(sheet_name)
        ws[f'{column_letter}{row}'] = value
        logger.debug(f"Cell '{column_letter}{row}' in '{sheet_name}' updated to: {value}")

    def get_current_quantity(self, sheet_name: str, row: int) -> int:
        """Gets the current quantity of an item.
        Assumes quantity is in column G for Ingresos/Salidas and C for Control."""
        if sheet_name in ['Ingresos de almacén', 'Salidas de almacén']:
            quantity_col = 'G'
        elif sheet_name == 'Control de inventarios':
            quantity_col = 'C'
        else:
            raise ValueError(f"Unknown sheet_name for quantity: {sheet_name}")

        quantity = self.get_cell_value(sheet_name, row, quantity_col)
        return quantity if quantity is not None else 0

    def get_max_row(self, sheet_name: str) -> int:
        """Returns the maximum row with data in a given sheet, considering column A."""
        ws = self.get_sheet(sheet_name)
        max_r = 2 # Start checking from row 3 (after headers)
        for row_idx in range(3, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value is not None: # Check if column A has data
                max_r = row_idx
        return max_r


# --- BaseTabManager Class ---
class BaseTabManager:
    """Base class for managing tabs with common functionalities."""

    def __init__(self, tab, excel_manager: ExcelManager, tab_name: str):
        self.tab = tab
        self.excel_manager = excel_manager
        self.tab_name = tab_name
        self.entries = {}
        self.setup_ui()

    def setup_ui(self):
        """Configures the base user interface."""
        raise NotImplementedError

    def create_form_field(self, label: str, row: int, column: int = 0, required: bool = True):
        """Creates a form field with label and entry."""
        display_label = f"{label} {'*' if required else ''}"
        tk.Label(self.tab, text=display_label).grid(
            row=row, column=column, sticky='e', padx=5, pady=5)
        entry = tk.Entry(self.tab, width=40)
        entry.grid(row=row, column=column + 1, padx=5, pady=5)
        self.entries[label] = entry
        return entry

    def validate_required_fields(self, fields: Dict[str, str]) -> bool:
        """Validates that required fields are completed."""
        for field, value in fields.items():
            if not value.strip():
                messagebox.showerror("Error de Validación", f"El campo '{field}' es obligatorio.")
                return False
        return True

    def validate_positive_integer(self, field_name: str, value: str) -> Tuple[bool, int]:
        """Validates that a value is a positive integer."""
        try:
            num = int(value)
            if num <= 0:
                messagebox.showerror("Error de Validación", f"{field_name} debe ser un número entero positivo.")
                return False, 0
            return True, num
        except ValueError:
            messagebox.showerror("Error de Validación", f"{field_name} debe ser un número entero válido.")
            return False, 0

    def clear_form(self):
        """Clears all form fields."""
        for entry in self.entries.values():
            entry.delete(0, tk.END)

# --- IngresoManager Class ---
class IngresoManager(BaseTabManager):
    """Handler for the 'Ingresos' tab."""

    def __init__(self, tab, excel_manager: ExcelManager):
        self.field_labels = [
            'N° de parte', 'Nombre', 'Descripción', 'Unidad',
            'Cantidad', 'Almacén', 'Ubicación', 'Encargado', 'Comentarios'
        ]
        super().__init__(tab, excel_manager, "Ingreso de artículos")

    def setup_ui(self):
        """Configures the user interface for income."""
        for i, label in enumerate(self.field_labels):
            is_required = label not in ['Comentarios', 'Descripción'] # Make Description optional too if needed
            self.create_form_field(label, i, required=is_required)

        tk.Button(
            self.tab,
            text="Guardar Ingreso",
            command=self.guardar_ingreso,
            bg="#4CAF50",
            fg="white",
            padx=10,
            pady=5,
            font=('Helvetica', 10, 'bold')
        ).grid(row=len(self.field_labels), column=0, columnspan=2, pady=10)

    def guardar_ingreso(self):
        """Handles the process of saving an income."""
        try:
            datos = {campo: entry.get().strip() for campo, entry in self.entries.items()}

            # Validate required fields (excluding comments and description)
            required_data = {k: v for k, v in datos.items() if k not in ['Comentarios', 'Descripción']}
            if not self.validate_required_fields(required_data):
                return

            valid, cantidad = self.validate_positive_integer("Cantidad", datos['Cantidad'])
            if not valid:
                return
            datos['Cantidad'] = cantidad # Update with validated integer

            # 1. Save or update to 'Ingresos de almacén'
            ws_ingresos = self.excel_manager.get_sheet('Ingresos de almacén')
            fila_existente_ingresos = self.excel_manager.find_part('Ingresos de almacén', datos['N° de parte'])

            if fila_existente_ingresos:
                self.actualizar_existente(fila_existente_ingresos, datos)
            else:
                self.crear_nuevo(datos)

            # 2. Then, update inventory control
            control_manager = ControlInventarioManager(self.excel_manager)
            control_manager.actualizar_inventario() # This will ensure min/max are read if they exist

            # 3. Finally, save all changes
            self.excel_manager.save()

            messagebox.showinfo("Éxito", "Ingreso registrado correctamente en ambas hojas.")
            self.clear_form()

        except Exception as e:
            logger.error(f"Error in guardar_ingreso: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Ocurrió un error al guardar el ingreso: {e}")

    def actualizar_existente(self, fila: int, datos: Dict):
        """Updates an existing record in 'Ingresos de almacén' sheet."""
        # Get current quantity in 'Ingresos de almacén'
        cantidad_actual_ingresos = self.excel_manager.get_cell_value('Ingresos de almacén', fila, 'G') or 0
        nueva_cantidad_ingresos = cantidad_actual_ingresos + datos['Cantidad']
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'G', nueva_cantidad_ingresos)

        # Update other fields for the existing entry if they are provided (overwriting)
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'A', datetime.now().strftime("%Y-%m-%d"))
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'C', datos['Nombre'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'D', datos['Descripción'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'F', datos['Unidad'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'H', datos['Almacén'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'I', datos['Ubicación'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'J', datos['Encargado'])
        self.excel_manager.update_cell('Ingresos de almacén', fila, 'K', datos['Comentarios'])
        logger.info(f"Existing part {datos['N° de parte']} updated in 'Ingresos de almacén' row {fila}. New quantity: {nueva_cantidad_ingresos}")

    def crear_nuevo(self, datos: Dict):
        """Creates a new record in 'Ingresos de almacén' sheet."""
        ws = self.excel_manager.get_sheet('Ingresos de almacén')
        next_row = self.excel_manager.get_max_row('Ingresos de almacén') + 1

        ws[f"A{next_row}"] = datetime.now().strftime("%Y-%m-%d")
        ws[f"B{next_row}"] = datos['N° de parte']
        ws[f"C{next_row}"] = datos['Nombre']
        ws[f"D{next_row}"] = datos['Descripción']
        ws[f"F{next_row}"] = datos['Unidad']
        ws[f"G{next_row}"] = datos['Cantidad']
        ws[f"H{next_row}"] = datos['Almacén']
        ws[f"I{next_row}"] = datos['Ubicación']
        ws[f"J{next_row}"] = datos['Encargado']
        ws[f"K{next_row}"] = datos['Comentarios']
        logger.info(f"New part {datos['N° de parte']} created in 'Ingresos de almacén' row {next_row}.")


# --- SalidaManager Class ---
class SalidaManager(BaseTabManager):
    """Handler for the 'Salidas' tab."""

    def __init__(self, tab, excel_manager: ExcelManager):
        self.field_labels = [
            'N° de parte', 'Nombre', 'Descripción', 'Unidad',
            'Cantidad', 'Almacén', 'Ubicación', 'Encargado', 'Comentarios'
        ]
        super().__init__(tab, excel_manager, "Salida de artículos")

    def setup_ui(self):
        """Configures the user interface for outputs."""
        for i, label in enumerate(self.field_labels):
            is_required = label not in ['Comentarios', 'Descripción']
            self.create_form_field(label, i, required=is_required)

        tk.Button(
            self.tab,
            text="Guardar Salida",
            command=self.guardar_salida,
            bg="#FF5722",
            fg="white",
            padx=10,
            pady=5,
            font=('Helvetica', 10, 'bold')
        ).grid(row=len(self.field_labels), column=0, columnspan=2, pady=10)

    def guardar_salida(self):
        """Handles the process of saving an output."""
        try:
            datos = {campo: entry.get().strip() for campo, entry in self.entries.items()}

            required_data = {k: v for k, v in datos.items() if k not in ['Comentarios', 'Descripción']}
            if not self.validate_required_fields(required_data):
                return

            valid, cantidad = self.validate_positive_integer("Cantidad", datos['Cantidad'])
            if not valid:
                return
            datos['Cantidad'] = cantidad

            # Verify existence in income (main inventory tracking)
            # We are checking 'Control de inventarios' for the stock status, but the
            # actual deduction of stock happens in 'Ingresos de almacén'.
            fila_ingreso = self.excel_manager.find_part('Ingresos de almacén', datos['N° de parte'])
            if not fila_ingreso:
                messagebox.showerror("Error", "El N° de parte no existe en el registro de ingresos. No se puede realizar la salida.")
                return

            # Verify available quantity from 'Ingresos de almacén'
            cantidad_disponible = self.excel_manager.get_current_quantity('Ingresos de almacén', fila_ingreso)
            if cantidad > cantidad_disponible:
                messagebox.showerror(
                    "Error",
                    f"Cantidad insuficiente. Disponible: {cantidad_disponible}, Solicitado: {cantidad}"
                )
                return

            # Register output in 'Salidas de almacén'
            self.registrar_salida(datos)

            # Update inventory (reduce quantity in 'Ingresos de almacén')
            nueva_cantidad = cantidad_disponible - cantidad
            self.excel_manager.update_cell('Ingresos de almacén', fila_ingreso, 'G', nueva_cantidad)
            logger.info(f"Stock for part {datos['N° de parte']} updated in 'Ingresos de almacén' to {nueva_cantidad}.")

            self.excel_manager.save()
            messagebox.showinfo("Éxito", "Salida registrada y cantidad actualizada correctamente.")
            self.clear_form()

            # Automatically update inventory control
            ControlInventarioManager(self.excel_manager).actualizar_inventario()

        except Exception as e:
            logger.error(f"Error in guardar_salida: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"Ocurrió un error al guardar la salida: {e}")

    def registrar_salida(self, datos: Dict):
        """Registers a new output in 'Salidas de almacén' Excel sheet."""
        ws = self.excel_manager.get_sheet('Salidas de almacén')
        next_row = self.excel_manager.get_max_row('Salidas de almacén') + 1

        ws[f"A{next_row}"] = datetime.now().strftime("%Y-%m-%d")
        ws[f"B{next_row}"] = datos['N° de parte']
        ws[f"C{next_row}"] = datos['Nombre']
        ws[f"D{next_row}"] = datos['Descripción']
        ws[f"F{next_row}"] = datos['Unidad']
        ws[f"G{next_row}"] = datos['Cantidad']
        ws[f"H{next_row}"] = datos['Almacén']
        ws[f"I{next_row}"] = datos['Ubicación']
        ws[f"J{next_row}"] = datos['Encargado']
        ws[f"K{next_row}"] = datos['Comentarios']
        logger.info(f"New part {datos['N° de parte']} output registered in 'Salidas de almacén' row {next_row}.")

# --- ConsultaManager Class ---
class ConsultaManager:
    """Handler for the 'Consultas' tab with separate visualization."""

    def __init__(self, tab, excel_manager: ExcelManager):
        self.tab = tab
        self.excel_manager = excel_manager
        self.setup_ui()

    def setup_ui(self):
        """Configures the user interface for queries with internal tabs."""
        # Button frame
        btn_frame = tk.Frame(self.tab)
        btn_frame.pack(fill='x', padx=10, pady=5)

        buttons = [
            ("Cargar Todo", "#9C27B0", self.cargar_todo),
            ("Cargar Ingresos", "#2196F3", lambda: self.cargar_datos('Ingresos de almacén')),
            ("Cargar Salidas", "#FF9800", lambda: self.cargar_datos('Salidas de almacén')),
            ("Ver Inventario", "#607D8B", self.mostrar_inventario)
        ]

        for text, color, command in buttons:
            tk.Button(
                btn_frame,
                text=text,
                command=command,
                bg=color,
                fg="white",
                padx=10,
                pady=5,
                font=('Helvetica', 9, 'bold')
            ).pack(side='left', padx=5)

        # Notebook for separate tables
        self.tabs_control = ttk.Notebook(self.tab)
        self.tabs_control.pack(expand=1, fill='both', padx=10, pady=10)

        # Create tabs for income, outcome, and inventory
        self.tab_ingresos = ttk.Frame(self.tabs_control)
        self.tab_salidas = ttk.Frame(self.tabs_control)
        self.tab_inventario = ttk.Frame(self.tabs_control)

        self.tabs_control.add(self.tab_ingresos, text="📥 Ingresos")
        self.tabs_control.add(self.tab_salidas, text="📤 Salidas")
        self.tabs_control.add(self.tab_inventario, text="📊 Inventario")

        # Treeviews for each tab
        self.tree_ingresos = self.create_treeview(self.tab_ingresos)
        self.tree_salidas = self.create_treeview(self.tab_salidas)
        self.tree_inventario = self.create_treeview(self.tab_inventario)

    def create_treeview(self, parent):
        """Creates a Treeview with scrollbars."""
        container = ttk.Frame(parent)
        container.pack(expand=True, fill='both')

        tree = ttk.Treeview(container)
        tree.pack(side='left', expand=True, fill='both')

        vsb = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        hsb.pack(side='bottom', fill='x')

        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        return tree

    def cargar_todo(self):
        """Loads and displays data from income, outcome, and inventory."""
        try:
            # Load income data
            df_ingresos = pd.read_excel(
                self.excel_manager.archivo_excel,
                sheet_name='Ingresos de almacén',
                skiprows=2 # Ensure this matches your header structure
            ).dropna(how='all') # Drop rows that are entirely NaN

            # Load outcome data
            df_salidas = pd.read_excel(
                self.excel_manager.archivo_excel,
                sheet_name='Salidas de almacén',
                skiprows=2 # Ensure this matches your header structure
            ).dropna(how='all') # Drop rows that are entirely NaN

            self.mostrar_datos(self.tree_ingresos, df_ingresos)
            self.mostrar_datos(self.tree_salidas, df_salidas)

            # Load inventory data
            self.mostrar_inventario()

            # Auto-adjust column width
            self.autoajustar_columnas(self.tree_ingresos)
            self.autoajustar_columnas(self.tree_salidas)
            self.autoajustar_columnas(self.tree_inventario)

            messagebox.showinfo("Éxito", "Datos cargados correctamente en las pestañas correspondientes.")

        except Exception as e:
            logger.error(f"Error loading all data: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"No se pudo cargar los datos: {e}")

    def mostrar_inventario(self):
        """Displays the current inventory status."""
        try:
            # Ensure inventory is updated before displaying
            ControlInventarioManager(self.excel_manager).actualizar_inventario()
            self.excel_manager.save() # Save changes made by actualizar_inventario

            df_inventario = pd.read_excel(
                self.excel_manager.archivo_excel,
                sheet_name='Control de inventarios',
                skiprows=2 # Ensure this matches your header structure
            ).dropna(how='all') # Drop rows that are entirely NaN

            # Configure specific columns for inventory
            self.tree_inventario["columns"] = list(df_inventario.columns)
            self.tree_inventario["show"] = "headings"

            for col in df_inventario.columns:
                self.tree_inventario.heading(col, text=col)
                self.tree_inventario.column(col, width=100, anchor='center', stretch=tk.YES)

            # Clear and load data
            self.tree_inventario.delete(*self.tree_inventario.get_children())

            for _, row in df_inventario.iterrows():
                values = list(row)
                estado = str(values[-1]) if len(values) > 5 else "" # Convert to string for 'in' check

                # Assign tags for coloring based on status
                tags = ()
                if "AGOTADO" in estado or "URGENTE" in estado:
                    tags = ('agotado',)
                elif "ALERTA" in estado:
                    tags = ('alerta',)
                elif "ADVERTENCIA" in estado:
                    tags = ('advertencia',)

                self.tree_inventario.insert("", "end", values=values, tags=tags)

            # Configure colors for tags
            self.tree_inventario.tag_configure('agotado', background='#ffcdd2')  # Light Red
            self.tree_inventario.tag_configure('alerta', background='#fff9c4')  # Light Yellow
            self.tree_inventario.tag_configure('advertencia', background='#ffcc80') # Light Orange

            # Select inventory tab
            self.tabs_control.select(2) # Index 2 is the 'Inventario' tab

        except Exception as e:
            logger.error(f"Error loading inventory: {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"No se pudo cargar el inventario: {e}")

    def mostrar_datos(self, tree: ttk.Treeview, df: pd.DataFrame):
        """Displays data in the specified Treeview."""
        # Clear treeview
        tree.delete(*tree.get_children())

        # Configure columns
        tree["columns"] = list(df.columns)
        tree["show"] = "headings"

        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor='center', stretch=tk.YES)

        # Insert data
        for _, row in df.iterrows():
            tree.insert("", "end", values=list(row))

    def autoajustar_columnas(self, tree: ttk.Treeview):
        """Automatically adjusts column width to content."""
        tree.update_idletasks()
        for col in tree["columns"]:
            # Consider both header and content length
            header_len = len(str(tree.heading(col)["text"])) * 8
            content_lens = [len(str(tree.set(item, col))) * 7 for item in tree.get_children()]
            max_len = max(header_len, *content_lens)
            tree.column(col, width=min(300, max_len + 10))  # Max limit of 300

    def cargar_datos(self, sheet_name: str):
        """Method to load data into the corresponding tab."""
        try:
            df = pd.read_excel(
                self.excel_manager.archivo_excel,
                sheet_name=sheet_name,
                skiprows=2 # Adjust as needed for your actual header rows
            ).dropna(how='all') # Drop rows that are entirely NaN

            target_tree = self.tree_ingresos if "Ingresos" in sheet_name else self.tree_salidas
            self.mostrar_datos(target_tree, df)
            self.autoajustar_columnas(target_tree)

            # Select the corresponding tab
            tab_index = 0 if "Ingresos" in sheet_name else 1
            self.tabs_control.select(tab_index)

            messagebox.showinfo("Éxito", f"Datos de '{sheet_name}' cargados correctamente.")

        except Exception as e:
            logger.error(f"Error loading '{sheet_name}': {str(e)}", exc_info=True)
            messagebox.showerror("Error", f"No se pudo cargar los datos de '{sheet_name}': {e}")

# --- ControlInventarioManager Class ---
class ControlInventarioManager:
    """Handler for advanced inventory control."""

    def __init__(self, excel_manager: ExcelManager):
        self.excel_manager = excel_manager
        self.umbral_alerta = 0.2  # 20% below minimum to alert

    def actualizar_inventario(self):
        """Updates the 'Control de Inventarios' sheet based on income and outcome."""
        try:
            wb = self.excel_manager.workbook
            ws_control = wb['Control de inventarios']
            ws_ingresos = wb['Ingresos de almacén']
            ws_salidas = wb['Salidas de almacén'] # Always assume this sheet exists

            # Create a dictionary to hold the aggregated stock and initial info for each part
            inventario_temp = {} # {part_number: {'nombre': '', 'stock_actual': 0, 'min': 0, 'max': 0}}

            # Aggregate stock from 'Ingresos de almacén'
            for row in range(3, ws_ingresos.max_row + 1):
                part = ws_ingresos.cell(row=row, column=2).value # Column B 'N° de parte'
                if part:
                    part = str(part).strip()
                    nombre = ws_ingresos.cell(row=row, column=3).value or "" # Column C 'Nombre'
                    cantidad = ws_ingresos.cell(row=row, column=7).value or 0 # Column G 'Cantidad'

                    if part not in inventario_temp:
                        inventario_temp[part] = {'nombre': nombre, 'stock_actual': 0, 'min': 0, 'max': 0}
                    inventario_temp[part]['stock_actual'] += cantidad
                    # Update name in case it's more complete in a later entry
                    if nombre:
                        inventario_temp[part]['nombre'] = nombre

            # Deduct stock from 'Salidas de almacén'
            for row in range(3, ws_salidas.max_row + 1):
                part = ws_salidas.cell(row=row, column=2).value # Column B 'N° de parte'
                if part:
                    part = str(part).strip()
                    cantidad = ws_salidas.cell(row=row, column=7).value or 0 # Column G 'Cantidad'
                    if part in inventario_temp:
                        inventario_temp[part]['stock_actual'] -= cantidad
                    else:
                        logger.warning(f"Part '{part}' found in 'Salidas' but not in 'Ingresos'. Skipping deduction for inventory control.")

            # Read existing min/max values from 'Control de inventarios'
            existing_control_data = {} # {part_number: {'min': value, 'max': value}}
            for row in range(3, ws_control.max_row + 1):
                part = ws_control.cell(row=row, column=1).value # Column A 'N° de parte'
                if part:
                    part = str(part).strip()
                    min_val = ws_control.cell(row=row, column=4).value # Column D 'Stock mínimo'
                    max_val = ws_control.cell(row=row, column=5).value # Column E 'Stock máximo'
                    existing_control_data[part] = {'min': min_val or 0, 'max': max_val or 0}

            # Clear existing data in 'Control de inventarios' (preserving headers)
            # Find the last actual row by checking for data in column A
            last_data_row = 2 # Start just after headers
            for r in range(3, ws_control.max_row + 1):
                if ws_control.cell(row=r, column=1).value is not None:
                    last_data_row = r
            # Clear from the first data row down
            for row in range(3, last_data_row + 1):
                for col_idx in range(1, 7): # Columns A to F
                    ws_control.cell(row=row, column=col_idx).value = None

            # Write updated data to 'Control de inventarios'
            next_row = 3
            for part, data in inventario_temp.items():
                stock_actual = max(0, data['stock_actual']) # Ensure stock doesn't go below 0

                # Use existing min/max if available, otherwise calculate defaults
                stock_minimo = existing_control_data.get(part, {}).get('min', 0)
                stock_maximo = existing_control_data.get(part, {}).get('max', 0)

                # If min/max are still 0 (i.e., not set manually or found in existing data), apply defaults
                if stock_minimo == 0 and stock_actual > 0: # Only calculate if stock exists
                    stock_minimo = max(int(stock_actual * 0.3), 1) # 30% of current stock as minimum
                if stock_maximo == 0 and stock_actual > 0: # Only calculate if stock exists
                    stock_maximo = max(int(stock_actual * 2), stock_minimo + 1) # Double current stock as maximum

                # If stock is 0 and no min/max exists, keep min/max at 0 to avoid large default numbers
                if stock_actual == 0 and stock_minimo == 0:
                    stock_minimo = 0
                if stock_actual == 0 and stock_maximo == 0:
                    stock_maximo = 0

                estado = self.determinar_estado(stock_actual, stock_minimo, stock_maximo)

                ws_control.cell(row=next_row, column=1, value=part) # A
                ws_control.cell(row=next_row, column=2, value=data['nombre']) # B
                ws_control.cell(row=next_row, column=3, value=stock_actual) # C
                ws_control.cell(row=next_row, column=4, value=stock_minimo) # D
                ws_control.cell(row=next_row, column=5, value=stock_maximo) # E
                ws_control.cell(row=next_row, column=6, value=estado) # F

                next_row += 1

            self.excel_manager.save()
            logger.info("Control de inventario updated successfully.")

        except Exception as e:
            logger.error(f"Error updating inventory: {str(e)}", exc_info=True)
            messagebox.showerror("Error de Inventario", f"Ocurrió un error al actualizar el inventario: {e}")
            raise

    def determinar_estado(self, actual: int, minimo: int, maximo: int) -> str:
        """Determines the inventory status with advanced logic."""
        if actual <= 0:
            return "🔴 AGOTADO - Sin stock disponible"
        elif actual <= minimo * (1 + self.umbral_alerta): # Below min + 20% threshold
            return "🟠 ALERTA - Stock por debajo del mínimo"
        elif actual <= minimo * 1.5: # 1.5 times min
            return "🟡 ADVERTENCIA - Stock próximo al mínimo"
        elif actual >= maximo * 0.9 and maximo > 0: # Nearing max stock, and max is set
            return "🟢 COMPLETO - Stock máximo alcanzado"
        else:
            return "⚪ NORMAL - Stock dentro de rangos"

    def predecir_necesidades(self, dias_historial: int = 30):
        """Predicts inventory needs based on historical data."""
        try:
            wb = self.excel_manager.workbook
            ws_control = wb['Control de inventarios']
            ws_salidas = wb['Salidas de almacén']

            fecha_limite = datetime.now() - timedelta(days=dias_historial)

            # Group outputs by part number within the historical period
            salidas_data = []
            for row in range(3, ws_salidas.max_row + 1):
                fecha_salida_raw = ws_salidas.cell(row=row, column=1).value
                # Ensure fecha_salida is a datetime object for comparison
                if isinstance(fecha_salida_raw, datetime):
                    fecha_salida = fecha_salida_raw
                elif isinstance(fecha_salida_raw, str):
                    try:
                        fecha_salida = datetime.strptime(fecha_salida_raw, "%Y-%m-%d")
                    except ValueError:
                        fecha_salida = None # Invalid date format
                else:
                    fecha_salida = None

                if fecha_salida and fecha_salida >= fecha_limite:
                    part = str(ws_salidas.cell(row=row, column=2).value).strip()
                    cantidad = ws_salidas.cell(row=row, column=7).value or 0
                    if part:
                        salidas_data.append({'N° de parte': part, 'Cantidad': cantidad})

            df_salidas_hist = pd.DataFrame(salidas_data)

            if not df_salidas_hist.empty:
                consumo_por_parte = df_salidas_hist.groupby('N° de parte')['Cantidad'].sum()
            else:
                consumo_por_parte = pd.Series(dtype=float)

            # Update control sheet based on predictions
            for row_idx in range(3, ws_control.max_row + 1):
                part = ws_control.cell(row=row_idx, column=1).value
                if part:
                    part = str(part).strip()
                    stock_actual = ws_control.cell(row=row_idx, column=3).value or 0 # Column C 'Stock actual'

                    consumo_total = consumo_por_parte.get(part, 0)

                    # Get current min/max to potentially update them
                    current_min = ws_control.cell(row=row_idx, column=4).value or 0
                    current_max = ws_control.cell(row=row_idx, column=5).value or 0

                    if consumo_total > 0 and dias_historial > 0:
                        consumo_diario = consumo_total / dias_historial
                        dias_restantes = stock_actual / consumo_diario if consumo_diario > 0 else float('inf')

                        suggested_min = max(int(consumo_diario * 15), 1) # Minimum for 15 days of consumption
                        suggested_max = int(consumo_diario * 30) # Maximum for 30 days of consumption

                        # Only update min/max if they are currently 0 or if the suggested value is higher
                        if current_min == 0 or suggested_min > current_min:
                             ws_control.cell(row=row_idx, column=4, value=suggested_min)
                        if current_max == 0 or suggested_max > current_max:
                             ws_control.cell(row=row_idx, column=5, value=suggested_max)

                        # Update status with prediction
                        if dias_restantes < 7:
                            estado = f"🔴 URGENTE - Solo {int(dias_restantes)} días de stock"
                        elif dias_restantes < 15:
                            estado = f"🟠 ALERTA - {int(dias_restantes)} días de stock"
                        else:
                            estado = "🟢 SUFICIENTE"
                        ws_control.cell(row=row_idx, column=6, value=estado) # Column F 'Estado'
                    else:
                         # If no consumption in history, re-evaluate status based on existing min/max
                         # Re-read min/max after potential suggestion updates above
                         updated_min = ws_control.cell(row=row_idx, column=4).value or 0
                         updated_max = ws_control.cell(row=row_idx, column=5).value or 0
                         ws_control.cell(row=row_idx, column=6, value=self.determinar_estado(stock_actual, updated_min, updated_max))


            self.excel_manager.save()
            logger.info(f"Prediction of needs completed for {dias_historial} days.")
            messagebox.showinfo("Predicción Completada", f"La predicción de necesidades se ha actualizado en la hoja 'Control de inventarios' (basado en {dias_historial} días de historial).")

        except Exception as e:
            logger.error(f"Error in predecir_necesidades: {str(e)}", exc_info=True)
            messagebox.showerror("Error de Predicción", f"Ocurrió un error al predecir necesidades: {e}")
            raise

    def generar_reporte(self):
        """Generates a complete inventory status report."""
        try:
            wb = self.excel_manager.workbook
            ws_control = wb['Control de inventarios']

            reporte = {
                "total_items": 0,
                "agotados": 0,
                "alertas": 0,
                "advertencias": 0,
                "sugerencias_reabastecimiento": []
            }

            for row in range(3, ws_control.max_row + 1):
                part = ws_control.cell(row=row, column=1).value
                if not part:
                    continue

                reporte["total_items"] += 1
                estado = ws_control.cell(row=row, column=6).value or ""

                if "AGOTADO" in estado or "URGENTE" in estado:
                    reporte["agotados"] += 1
                elif "ALERTA" in estado:
                    reporte["alertas"] += 1
                elif "ADVERTENCIA" in estado:
                    reporte["advertencias"] += 1

                stock_actual = ws_control.cell(row=row, column=3).value or 0
                stock_minimo = ws_control.cell(row=row, column=4).value or 0

                if stock_actual < stock_minimo:
                    reporte["sugerencias_reabastecimiento"].append({
                        "parte": part,
                        "nombre": ws_control.cell(row=row, column=2).value,
                        "actual": stock_actual,
                        "minimo": stock_minimo,
                        "cantidad_sugerida": max(stock_minimo - stock_actual + 1, 1) # Suggest to reach at least min + 1
                    })

            # Display the report in a message box or new window
            report_str = f"""--- Reporte de Inventario ---
Total de ítems: {reporte['total_items']}
Ítems agotados/urgentes: {reporte['agotados']}
Ítems en alerta: {reporte['alertas']}
Ítems en advertencia: {reporte['advertencias']}

Sugerencias de Reabastecimiento:
"""
            if reporte["sugerencias_reabastecimiento"]:
                for item in reporte["sugerencias_reabastecimiento"]:
                    report_str += (f"- N° Parte: {item['parte']}, Nombre: {item['nombre']}, "
                                   f"Actual: {item['actual']}, Mínimo: {item['minimo']}, "
                                   f"Sugerido: {item['cantidad_sugerida']}\n")
            else:
                report_str += "Ninguna sugerencia de reabastecimiento en este momento."

            messagebox.showinfo("Reporte de Inventario", report_str)
            logger.info("Inventory report generated.")

            return reporte

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}", exc_info=True)
            messagebox.showerror("Error de Reporte", f"Ocurrió un error al generar el reporte: {e}")
            raise

# --- Main Application Setup ---
def crear_pestanas(root, archivo_excel: str):
    """Main function to create the tabs."""
    instructions = (
        "\nSistema de Gestión en Almacén de Equipos y Herramientas\n"
        "📌 Ingreso: Complete campos obligatorios (*) y presione 'Guardar Ingreso'\n"
        "📌 Salida: Complete campos obligatorios (*) y presione 'Guardar Salida'\n"
        "📌 Consulta: Use los botones para visualizar ingresos, salidas e inventario\n"
    )

    # Frame for instructions with better style
    instruction_frame = tk.Frame(root, bd=1, relief=tk.GROOVE, bg="#E0F7FA") # Light blue background
    instruction_frame.pack(fill="x", padx=10, pady=5)

    label_info = tk.Label(
        instruction_frame,
        text=instructions,
        justify="left",
        anchor="w",
        padx=10,
        pady=10,
        font=('Helvetica', 10, 'italic'), # Italic font
        bg="#E0F7FA",
        fg="#006064" # Darker blue text
    )
    label_info.pack(fill="x")

    # Create notebook (tabs)
    tabControl = ttk.Notebook(root)

    tabs = {
        'ingreso': ttk.Frame(tabControl),
        'salida': ttk.Frame(tabControl),
        'consulta': ttk.Frame(tabControl)
    }

    tabControl.add(tabs['ingreso'], text='📥 Ingreso de artículos')
    tabControl.add(tabs['salida'], text='📤 Salida de artículos')
    tabControl.add(tabs['consulta'], text='🔍 Consulta de registros')
    tabControl.pack(expand=1, fill="both", padx=10, pady=10)

    # Initialize ExcelManager and tab managers
    excel_manager = ExcelManager(archivo_excel)

    IngresoManager(tabs['ingreso'], excel_manager)
    SalidaManager(tabs['salida'], excel_manager)
    consulta_manager = ConsultaManager(tabs['consulta'], excel_manager) # Keep a reference if needed later

    # Add buttons for advanced inventory control
    advanced_btn_frame = tk.Frame(root)
    advanced_btn_frame.pack(pady=5)

    tk.Button(
        advanced_btn_frame,
        text="🔄 Predecir Necesidades",
        command=lambda: ControlInventarioManager(excel_manager).predecir_necesidades(),
        bg="#00796B", # Teal
        fg="white",
        padx=10,
        pady=5,
        font=('Helvetica', 9, 'bold')
    ).pack(side='left', padx=5)

    tk.Button(
        advanced_btn_frame,
        text="📋 Generar Reporte",
        command=lambda: ControlInventarioManager(excel_manager).generar_reporte(),
        bg="#5D4037", # Brown
        fg="white",
        padx=10,
        pady=5,
        font=('Helvetica', 9, 'bold')
    ).pack(side='left', padx=5)


    # Exit button with better style
    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=10)

    tk.Button(
        btn_frame,
        text="🚪 Salir",
        command=root.quit,
        bg="#f44336",
        fg="white",
        padx=20,
        pady=5,
        font=('Helvetica', 10, 'bold')
    ).pack()

def main():
    root = tk.Tk()
    root.title("Sistema de Gestión en Almacén de Equipos y Herramientas")
    root.geometry("1100x700")

    style = ttk.Style()
    style.theme_use('clam')
    style.configure('TNotebook.Tab', font=('Helvetica', 10, 'bold'), padding=[10, 5])
    style.configure('TFrame', background='#ECEFF1')
    style.configure('TButton', font=('Helvetica', 9))

    # Determine the base path for resources
    if getattr(sys, 'frozen', False):
        # If running as a bundled executable, use the temp folder created by PyInstaller
        base_path = sys._MEIPASS
    else:
        # If running as a script, use the current directory
        base_path = os.path.dirname(__file__)

    # Construct the path to the data directory and Excel file
    data_dir = os.path.join(base_path, 'data')
    default_excel_path = os.path.join(data_dir, 'Sistema de Gestion en Almacén de Equipos y Herramientas.xlsx')

    # Create the 'data' directory if it doesn't exist (important for dev and initial run)
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
        logger.info(f"Created data directory: {data_dir}")

    # Check if the default excel file exists, if not, create a blank one
    if not os.path.exists(default_excel_path):
        try:
            from openpyxl import Workbook
            wb_new = Workbook()
            if 'Sheet' in wb_new.sheetnames:
                del wb_new['Sheet']

            ws_ingresos = wb_new.create_sheet('Ingresos de almacén')
            ws_ingresos['A1'] = "REGISTRO DE INGRESOS EN ALMACÉN"
            ws_ingresos['A2'] = "Fecha"
            ws_ingresos['B2'] = "N° de parte"
            ws_ingresos['C2'] = "Nombre"
            ws_ingresos['D2'] = "Descripción"
            ws_ingresos['E2'] = " "
            ws_ingresos['F2'] = "Unidad"
            ws_ingresos['G2'] = "Cantidad"
            ws_ingresos['H2'] = "Almacén"
            ws_ingresos['I2'] = "Ubicación"
            ws_ingresos['J2'] = "Encargado"
            ws_ingresos['K2'] = "Comentarios"

            ws_salidas = wb_new.create_sheet('Salidas de almacén')
            ws_salidas['A1'] = "REGISTRO DE SALIDAS EN ALMACÉN"
            ws_salidas['A2'] = "Fecha"
            ws_salidas['B2'] = "N° de parte"
            ws_salidas['C2'] = "Nombre"
            ws_salidas['D2'] = "Descripción"
            ws_salidas['E2'] = " "
            ws_salidas['F2'] = "Unidad"
            ws_salidas['G2'] = "Cantidad"
            ws_salidas['H2'] = "Almacén"
            ws_salidas['I2'] = "Ubicación"
            ws_salidas['J2'] = "Encargado"
            ws_salidas['K2'] = "Comentarios"

            ws_control = wb_new.create_sheet('Control de inventarios')
            ws_control['A1'] = "CONTROL DE INVENTARIOS"
            ws_control['A2'] = "N° de parte"
            ws_control['B2'] = "Nombre"
            ws_control['C2'] = "Stock actual"
            ws_control['D2'] = "Stock mínimo"
            ws_control['E2'] = "Stock máximo"
            ws_control['F2'] = "Estado"

            wb_new.save(default_excel_path)
            logger.info(f"Created new blank Excel file at: {default_excel_path}")
            messagebox.showinfo("Archivo Excel Creado", f"Se ha creado un nuevo archivo Excel en:\n{default_excel_path}\nPor favor, configure los encabezados en las hojas 'Ingresos de almacén', 'Salidas de almacén' y 'Control de inventarios' si desea personalizarlos más allá de los valores predeterminados.")
        except Exception as e:
            logger.error(f"Error creating new Excel file: {str(e)}")
            messagebox.showerror("Error", f"No se pudo crear el archivo Excel predeterminado: {e}")
            return

    archivo_excel = default_excel_path

    crear_pestanas(root, archivo_excel)
    root.mainloop()

if __name__ == "__main__":
    main()